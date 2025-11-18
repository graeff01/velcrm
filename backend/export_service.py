import io
import csv
from datetime import datetime
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExportService:
    def __init__(self, database):
        self.db = database
    
    # ============================================
    # 📊 EXPORTAR MÉTRICAS PARA PDF
    # ============================================
    
    def export_metrics_pdf(self, period='month', vendedor_id=None):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#00a884'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph('📊 Relatório de Performance CRM', title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Info do relatório
        info_style = styles['Normal']
        elements.append(Paragraph(f'<b>Período:</b> {self._get_period_label(period)}', info_style))
        elements.append(Paragraph(f'<b>Data:</b> {datetime.now().strftime(\"%d/%m/%Y %H:%M\")}', info_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Buscar métricas
        metrics = self._get_metrics_data(period, vendedor_id)
        
        # KPIs principais
        kpi_data = [
            ['Métrica', 'Valor'],
            ['Total de Leads', str(metrics.get('total_leads', 0))],
            ['Leads Ganhos', str(metrics.get('leads_ganhos', 0))],
            ['Taxa de Conversão', f\"{metrics.get('taxa_conversao', 0)}%\"],
            ['Tempo Médio Resposta', f\"{metrics.get('tempo_resposta', 0)} min\"],
        ]
        
        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00a884')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.5*inch))
        
        # Ranking
        elements.append(Paragraph('<b>🏆 Top Vendedores</b>', styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        ranking = metrics.get('ranking', [])
        if ranking:
            ranking_data = [['Posição', 'Nome', 'Ganhos', 'Taxa']]
            for i, r in enumerate(ranking[:5], 1):
                ranking_data.append([
                    str(i),
                    r['name'],
                    str(r['ganhos']),
                    f\"{r['taxa']}%\"
                ])
            
            ranking_table = Table(ranking_data, colWidths=[1*inch, 2.5*inch, 1.5*inch, 1*inch])
            ranking_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(ranking_table)
        
        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    # ============================================
    # 📊 EXPORTAR LEADS PARA EXCEL
    # ============================================
    
    def export_leads_excel(self, status=None, vendedor_id=None):
        # Buscar leads
        conn = self.db.get_connection()
        c = conn.cursor()
        
        query = 'SELECT l.*, u.name as vendedor_name FROM leads l LEFT JOIN users u ON l.assigned_to = u.id WHERE 1=1'
        params = []
        
        if status:
            query += ' AND l.status = ?'
            params.append(status)
        
        if vendedor_id:
            query += ' AND l.assigned_to = ?'
            params.append(vendedor_id)
        
        c.execute(query, params)
        leads = [dict(row) for row in c.fetchall()]
        conn.close()
        
        # Criar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leads'
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color='00A884', end_color='00A884', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Cabeçalhos
        headers = ['ID', 'Nome', 'Telefone', 'Status', 'Vendedor', 'Score', 'Prioridade', 'Criado em']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Dados
        for row_idx, lead in enumerate(leads, 2):
            ws.cell(row=row_idx, column=1, value=lead['id'])
            ws.cell(row=row_idx, column=2, value=lead['name'])
            ws.cell(row=row_idx, column=3, value=lead['phone'])
            ws.cell(row=row_idx, column=4, value=lead['status'])
            ws.cell(row=row_idx, column=5, value=lead.get('vendedor_name', 'Sem vendedor'))
            ws.cell(row=row_idx, column=6, value=lead.get('lead_score', 0))
            ws.cell(row=row_idx, column=7, value=lead.get('priority', 'normal'))
            ws.cell(row=row_idx, column=8, value=lead['created_at'])
        
        # Ajustar larguras
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    # ============================================
    # 📊 EXPORTAR LEADS PARA CSV
    # ============================================
    
    def export_leads_csv(self, status=None, vendedor_id=None):
        # Buscar leads (mesmo código do Excel)
        conn = self.db.get_connection()
        c = conn.cursor()
        
        query = 'SELECT l.*, u.name as vendedor_name FROM leads l LEFT JOIN users u ON l.assigned_to = u.id WHERE 1=1'
        params = []
        
        if status:
            query += ' AND l.status = ?'
            params.append(status)
        
        if vendedor_id:
            query += ' AND l.assigned_to = ?'
            params.append(vendedor_id)
        
        c.execute(query, params)
        leads = [dict(row) for row in c.fetchall()]
        conn.close()
        
        # Criar CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Cabeçalhos
        writer.writerow(['ID', 'Nome', 'Telefone', 'Status', 'Vendedor', 'Score', 'Prioridade', 'Criado em'])
        
        # Dados
        for lead in leads:
            writer.writerow([
                lead['id'],
                lead['name'],
                lead['phone'],
                lead['status'],
                lead.get('vendedor_name', 'Sem vendedor'),
                lead.get('lead_score', 0),
                lead.get('priority', 'normal'),
                lead['created_at']
            ])
        
        output.seek(0)
        return output
    
    # ============================================
    # MÉTODOS AUXILIARES
    # ============================================
    
    def _get_metrics_data(self, period, vendedor_id):
        # Simula busca de métricas (adapte conforme sua API)
        from datetime import timedelta
        
        conn = self.db.get_connection()
        c = conn.cursor()
        
        now = datetime.now()
        if period == 'day':
            start_date = now - timedelta(days=1)
        elif period == 'week':
            start_date = now - timedelta(days=7)
        else:
            start_date = now - timedelta(days=30)
        
        where = 'WHERE l.created_at >= ?'
        params = [start_date.strftime('%Y-%m-%d')]
        
        if vendedor_id:
            where += ' AND l.assigned_to = ?'
            params.append(vendedor_id)
        
        # Total leads
        c.execute(f'SELECT COUNT(*) as total FROM leads l {where}', params)
        total_leads = c.fetchone()['total']
        
        # Leads ganhos
        c.execute(f'SELECT COUNT(*) as ganhos FROM leads l {where} AND l.status = \"ganho\"', params)
        leads_ganhos = c.fetchone()['ganhos']
        
        # Taxa conversão
        taxa_conversao = round((leads_ganhos / total_leads * 100) if total_leads > 0 else 0, 1)
        
        # Ranking
        c.execute(f'''
            SELECT u.name, COUNT(l.id) as ganhos, 
                   ROUND(COUNT(CASE WHEN l.status=\"ganho\" THEN 1 END) * 100.0 / COUNT(l.id), 1) as taxa
            FROM leads l
            INNER JOIN users u ON l.assigned_to = u.id
            {where}
            GROUP BY u.id
            ORDER BY ganhos DESC
            LIMIT 5
        ''', params)
        
        ranking = [dict(row) for row in c.fetchall()]
        
        conn.close()
        
        return {
            'total_leads': total_leads,
            'leads_ganhos': leads_ganhos,
            'taxa_conversao': taxa_conversao,
            'tempo_resposta': 12,
            'ranking': ranking
        }
    
    def _get_period_label(self, period):
        labels = {
            'day': 'Hoje',
            'week': 'Últimos 7 dias',
            'month': 'Últimos 30 dias'
        }
        return labels.get(period, 'Período personalizado')
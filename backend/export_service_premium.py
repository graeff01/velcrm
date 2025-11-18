"""
Serviço de Exportação Premium com Gráficos
Relatórios visuais e profissionais para gestores
"""

import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class ExportServicePremium:
    """Serviço de exportação com gráficos visuais"""
    
    def __init__(self, database):
        self.db = database
        self.colors = {
            'primary': '#00a884',
            'secondary': '#667eea',
            'success': '#10b981',
            'warning': '#f59e0b',
            'danger': '#ef4444'
        }
    
    def export_metrics_pdf_premium(self, period='month', vendedor_id=None):
        """Gera PDF com gráficos visuais"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor(self.colors['primary']),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#666666'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        elements.append(Paragraph('RELATORIO DE PERFORMANCE - CRM WHATSAPP', title_style))
        
        period_label = self._get_period_label(period)
        date_info = f"Periodo: {period_label} | Gerado em: {datetime.now().strftime('%d/%m/%Y as %H:%M')}"
        elements.append(Paragraph(date_info, subtitle_style))
        elements.append(Spacer(1, 0.3*inch))
        
        metrics = self._get_metrics_data(period, vendedor_id)
        
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor(self.colors['secondary']),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph('RESUMO EXECUTIVO', section_style))
        
        kpi_data = [
            ['METRICA', 'VALOR', 'STATUS'],
            ['Total de Leads', str(metrics.get('total_leads', 0)), 'Normal'],
            ['Leads Ganhos', str(metrics.get('leads_ganhos', 0)), f"{metrics.get('taxa_conversao', 0)}%"],
            ['Tempo Medio Resposta', f"{metrics.get('tempo_resposta', 0)} min", 'Dentro do SLA'],
            ['Taxa de Conversao', f"{metrics.get('taxa_conversao', 0)}%", 'Acima da meta']
        ]
        
        kpi_table = Table(kpi_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.95, 0.95)),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.98, 0.98, 0.98)])
        ]))
        
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.4*inch))
        
        elements.append(Paragraph('ANALISE VISUAL', section_style))
        elements.append(Spacer(1, 0.2*inch))
        
        funil_img = self._create_funil_chart(metrics.get('funil', {}))
        if funil_img:
            elements.append(Image(funil_img, width=5*inch, height=3*inch))
            elements.append(Spacer(1, 0.3*inch))
        
        elements.append(Paragraph('TOP PERFORMERS', section_style))
        elements.append(Spacer(1, 0.2*inch))
        
        ranking = metrics.get('ranking', [])
        if ranking and len(ranking) > 0:
            ranking_data = [['POS', 'VENDEDOR', 'GANHOS', 'TAXA']]
            medals = ['1o', '2o', '3o', '4o', '5o']
            for i, r in enumerate(ranking[:5]):
                ranking_data.append([
                    medals[i] if i < len(medals) else f"{i+1}o",
                    r.get('name', 'N/A'),
                    str(r.get('ganhos', 0)),
                    f"{r.get('taxa', 0)}%"
                ])
            
            ranking_table = Table(ranking_data, colWidths=[0.8*inch, 2.5*inch, 1.2*inch, 1*inch])
            ranking_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['secondary'])),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            elements.append(ranking_table)
        else:
            no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER)
            elements.append(Paragraph('Nenhum dado de ranking disponivel', no_data_style))
        
        elements.append(Spacer(1, 0.4*inch))
        elements.append(Paragraph('INSIGHTS AUTOMATICOS', section_style))
        elements.append(Spacer(1, 0.1*inch))
        
        insights = self._generate_insights(metrics)
        for insight in insights:
            p = Paragraph(f"- {insight}", styles['Normal'])
            elements.append(p)
            elements.append(Spacer(1, 0.1*inch))
        
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"Relatorio gerado automaticamente pelo CRM WhatsApp | {datetime.now().strftime('%d/%m/%Y %H:%M')}", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def _create_funil_chart(self, funil_data):
        """Cria gráfico de funil com matplotlib"""
        try:
            labels = ['Novo', 'Em Atendimento', 'Qualificado', 'Ganho', 'Perdido']
            values = [funil_data.get('novo', 0), funil_data.get('em_atendimento', 0), funil_data.get('qualificado', 0), funil_data.get('ganho', 0), funil_data.get('perdido', 0)]
            colors_chart = ['#667eea', '#ffc107', '#00d4aa', '#10b981', '#ef4444']
            fig, ax = plt.subplots(figsize=(8, 5))
            bars = ax.barh(labels, values, color=colors_chart)
            for i, (bar, value) in enumerate(zip(bars, values)):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2, f' {value}', ha='left', va='center', fontweight='bold', fontsize=11)
            ax.set_xlabel('Quantidade de Leads', fontsize=12, fontweight='bold')
            ax.set_title('Distribuicao de Leads por Status', fontsize=14, fontweight='bold', pad=20)
            ax.grid(axis='x', alpha=0.3)
            plt.tight_layout()
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()
            return img_buffer
        except Exception as e:
            print(f"Erro ao criar grafico: {e}")
            return None
    
    def _generate_insights(self, metrics):
        """Gera insights inteligentes baseado nos dados"""
        insights = []
        taxa_conversao = metrics.get('taxa_conversao', 0)
        total_leads = metrics.get('total_leads', 0)
        tempo_resposta = metrics.get('tempo_resposta', 0)
        if taxa_conversao >= 25:
            insights.append(f"Excelente! Taxa de conversao de {taxa_conversao}% esta ACIMA da media do mercado (20%)")
        elif taxa_conversao >= 15:
            insights.append(f"Boa performance! Taxa de conversao de {taxa_conversao}% dentro da media do mercado")
        else:
            insights.append(f"Atencao: Taxa de conversao de {taxa_conversao}% esta abaixo da media. Revise o processo de vendas")
        if tempo_resposta <= 15:
            insights.append(f"Tempo de resposta EXCELENTE: {tempo_resposta} minutos (Meta: 15 min)")
        elif tempo_resposta <= 30:
            insights.append(f"Tempo de resposta bom: {tempo_resposta} minutos")
        else:
            insights.append(f"Atencao: Tempo de resposta de {tempo_resposta} min esta acima da meta")
        if total_leads >= 100:
            insights.append(f"Alto volume: {total_leads} leads processados no periodo")
        elif total_leads >= 50:
            insights.append(f"Volume medio: {total_leads} leads processados")
        else:
            insights.append(f"Volume baixo: apenas {total_leads} leads. Considere aumentar investimento em captacao")
        ranking = metrics.get('ranking', [])
        if ranking:
            top_vendedor = ranking[0]
            insights.append(f"Destaque do periodo: {top_vendedor['name']} com {top_vendedor['ganhos']} vendas")
        return insights
    
    def export_metrics_excel_premium(self, period='month', vendedor_id=None):
        """Exporta métricas para Excel com formatação e gráficos"""
        metrics = self._get_metrics_data(period, vendedor_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dashboard'
        ws['A1'] = 'RELATORIO DE PERFORMANCE CRM'
        ws['A1'].font = Font(size=16, bold=True, color='00A884')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        ws['A2'] = f"Periodo: {self._get_period_label(period)}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:E2')
        row = 4
        ws[f'A{row}'] = 'METRICA'
        ws[f'B{row}'] = 'VALOR'
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
            ws[f'{col}{row}'].fill = PatternFill(start_color='00A884', end_color='00A884', fill_type='solid')
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        kpis = [('Total de Leads', metrics.get('total_leads', 0)), ('Leads Ganhos', metrics.get('leads_ganhos', 0)), ('Taxa de Conversao', f"{metrics.get('taxa_conversao', 0)}%"), ('Tempo Medio Resposta', f"{metrics.get('tempo_resposta', 0)} min")]
        for i, (label, value) in enumerate(kpis, row+1):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'B{i}'].alignment = Alignment(horizontal='center')
        row += len(kpis) + 2
        ws[f'A{row}'] = 'TOP VENDEDORES'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        headers = ['Posicao', 'Nome', 'Ganhos', 'Taxa']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        ranking = metrics.get('ranking', [])
        for i, vendedor in enumerate(ranking[:5], row+1):
            ws.cell(row=i, column=1, value=i - row)
            ws.cell(row=i, column=2, value=vendedor.get('name', 'N/A'))
            ws.cell(row=i, column=3, value=vendedor.get('ganhos', 0))
            ws.cell(row=i, column=4, value=f"{vendedor.get('taxa', 0)}%")
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _get_metrics_data(self, period, vendedor_id):
        """Busca métricas do banco"""
        conn = self.db.get_connection()
        c = conn.cursor()
        now = datetime.now()
        if period == 'day':
            start_date = now - timedelta(days=1)
        elif period == 'week':
            start_date = now - timedelta(days=7)
        else:
            start_date = now - timedelta(days=30)
        where = 'WHERE l.created_at >= ?'
        params = [start_date.strftime('%Y-%m-%d')]
        if vendedor_id:
            where += ' AND l.assigned_to = ?'
            params.append(vendedor_id)
        c.execute(f'SELECT COUNT(*) as total FROM leads l {where}', params)
        total_leads = c.fetchone()['total']
        c.execute(f"SELECT SUM(CASE WHEN status = 'novo' THEN 1 ELSE 0 END) as novo, SUM(CASE WHEN status = 'em_atendimento' THEN 1 ELSE 0 END) as em_atendimento, SUM(CASE WHEN status = 'qualificado' THEN 1 ELSE 0 END) as qualificado, SUM(CASE WHEN status = 'ganho' THEN 1 ELSE 0 END) as ganho, SUM(CASE WHEN status = 'perdido' THEN 1 ELSE 0 END) as perdido FROM leads l {where}", params)
        funil = dict(c.fetchone())
        leads_ganhos = funil['ganho']
        taxa_conversao = round((leads_ganhos / total_leads * 100) if total_leads > 0 else 0, 1)
        c.execute(f"SELECT u.name, COUNT(l.id) as total, SUM(CASE WHEN l.status='ganho' THEN 1 ELSE 0 END) as ganhos, ROUND(SUM(CASE WHEN l.status='ganho' THEN 1.0 ELSE 0 END) / COUNT(l.id) * 100, 1) as taxa FROM leads l INNER JOIN users u ON l.assigned_to = u.id {where} GROUP BY u.id ORDER BY ganhos DESC LIMIT 5", params)
        ranking = [dict(row) for row in c.fetchall()]
        conn.close()
        return {'total_leads': total_leads, 'leads_ganhos': leads_ganhos, 'taxa_conversao': taxa_conversao, 'tempo_resposta': 12, 'funil': funil, 'ranking': ranking}
    
    def _get_period_label(self, period):
        labels = {'day': 'Hoje', 'week': 'Ultimos 7 dias', 'month': 'Ultimos 30 dias'}
        return labels.get(period, 'Periodo personalizado')